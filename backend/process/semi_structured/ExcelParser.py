import os
from typing import Dict, Any
from .SemiStructuredProcessor import SemiStructuredProcessor


class ExcelParser(SemiStructuredProcessor):
    """
    Excel 文件解析器 (.xlsx)
    提取所有工作表的表格数据，使用 LLM 生成描述
    """

    type = "excel"

    def __init__(self, file_path: str, **kwargs):
        super().__init__(file_path, **kwargs)
        self._sheet_names: list = []
        self._row_count: int = 0

    def _extract_content(self) -> str:
        return self._extract_with_openpyxl()

    def _extract_with_openpyxl(self) -> str:
        """使用 openpyxl 提取 .xlsx / .ods 内容"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("⚠️ 未安装 openpyxl，尝试 pandas 作为降级方案")
            return self._extract_with_pandas()

        try:
            wb = load_workbook(self.file_path, read_only=True)
            text = []
            self._sheet_names = wb.sheetnames

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text.append(f"=== Sheet: {sheet_name} ===")

                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in row_text if cell):
                        text.append("\t".join(row_text))
                        self._row_count += 1

                text.append("")  # sheet 之间空行

            wb.close()
            extracted = "\n".join(text)
            return extracted if extracted.strip() else ""

        except Exception as e:
            print(f"❌ openpyxl 读取失败 {self.file_path}: {e}")
            return self._extract_with_pandas()

    def _extract_with_pandas(self) -> str:
        """使用 pandas 作为降级方案（支持 .xls）"""
        try:
            import pandas as pd
        except ImportError:
            print(f"❌ pandas 也不可用，无法解析 {self.file_path}")
            return ""

        try:
            xls = pd.ExcelFile(self.file_path)
            text = []
            self._sheet_names = xls.sheet_names

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                text.append(f"=== Sheet: {sheet_name} ===")
                text.append("\t".join(df.columns.astype(str)))

                for _, row in df.iterrows():
                    row_text = [str(v) if pd.notna(v) else "" for v in row]
                    if any(v.strip() for v in row_text if v):
                        text.append("\t".join(row_text))
                        self._row_count += 1

                text.append("")

            extracted = "\n".join(text)
            return extracted if extracted.strip() else ""

        except Exception as e:
            print(f"❌ pandas 读取 Excel 失败 {self.file_path}: {e}")
            return ""

    def _get_description_prompt(self, content_preview: str) -> str:
        file_name = os.path.basename(self.file_path)
        sheet_info = f"工作表: {', '.join(self._sheet_names[:5])}"

        return f"""根据以下Excel表格数据，生成一段用于语义检索的描述文本。

要求：
- 只输出描述文本本身，不要有任何前缀、解释或格式标记
- 说明这个表是什么内容（如：员工信息表、销售数据表）
- 包含表格中的关键名称、字段名等便于检索的信息
- 提及数据涉及的时间范围或地域（如果有）
- 根据内容多少灵活调整长度：简单表格简短描述，包含多个工作表的复杂文件可详细描述

文件名：{file_name}
{sheet_info}

表格数据：
{content_preview}"""

    def _get_fallback_description(self) -> str:
        file_name = os.path.basename(self.file_path)
        parts = [f"Excel表格 {file_name}"]
        if self._sheet_names:
            parts.append(f"工作表: {', '.join(self._sheet_names[:3])}")
        if self._content:
            preview = self._content[:300].replace("\n", " ")
            parts.append(preview)
        return " ".join(parts)

    @property
    def metadata(self) -> Dict[str, Any]:
        return {
            **super().metadata,
            "sheet_names": self._sheet_names,
            "row_count": self._row_count,
            "parser": "ExcelParser",
        }
